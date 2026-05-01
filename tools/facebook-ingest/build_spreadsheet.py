#!/usr/bin/env python3
"""
Build a color-tiered Excel spreadsheet of top Facebook contacts.

NOTE: the `contacts` list below is DUMMY sample data. Replace it with your
own top-N contacts (generated from analyze_contacts.py output).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Key Contacts"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
TIER1_FILL = PatternFill("solid", fgColor="D6E4F0")
TIER2_FILL = PatternFill("solid", fgColor="E2EFDA")
TIER3_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

headers = [
    "Rank", "Name", "Tier", "Score", "Messages", "Tags",
    "Active Period", "LinkedIn", "Twitter/X", "Instagram",
    "Facebook", "Website", "Current Role/Company", "Context Snippet"
]

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = HEADER_ALIGN
    c.border = THIN_BORDER

col_widths = [6, 28, 8, 7, 9, 32, 24, 38, 30, 30, 30, 30, 40, 55]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.auto_filter.ref = f"A1:N1"
ws.freeze_panes = "A2"

# ──────────────────────────────────────────────────────────────────────
# DUMMY DATA — replace with real contacts from your own analysis
# Columns: (name, tier, score, msgs, tags, active, linkedin, twitter,
#           instagram, facebook, website, role, snippet)
# ──────────────────────────────────────────────────────────────────────
contacts = [
    ("Alex Sample", 1, 95, 30, "Industry, Advisor, Professional",
     "Jan 2020 – Dec 2021", "linkedin.com/in/alex-sample", "@alexsample",
     "", "", "alexsample.example",
     "Founder, SampleCo", "Sample snippet — shared projects, frequent DMs"),
    ("Jordan Doe", 1, 90, 25, "Collaborator, Personal",
     "Mar 2019 – Jul 2021", "", "@jordandoe", "", "",
     "", "Designer, SampleStudio", "Sample snippet — met at conference"),
    ("Riley Example", 2, 80, 18, "Professional, Recruiting",
     "Jun 2018 – Feb 2020", "linkedin.com/in/riley-example", "",
     "", "", "", "PM, SampleApp",
     "Sample snippet — ongoing discussions"),
    ("Casey Placeholder", 2, 75, 15, "Personal, Connector",
     "2017 – 2019", "", "", "instagram.com/caseyplaceholder", "",
     "", "Artist / independent", "Sample snippet — studio visits"),
    ("Morgan Demo", 3, 65, 8, "Professional",
     "Sept 2016 – Dec 2017", "", "", "", "", "", "Founder, SampleLLC",
     "Sample snippet — early-stage collab"),
]

for i, row in enumerate(contacts):
    r = i + 2
    name, tier, score, msgs, tags, active, li, tw, ig, fb, web, role, snippet = row
    tier_label = f"Tier {tier}"
    fill = {1: TIER1_FILL, 2: TIER2_FILL, 3: TIER3_FILL}[tier]

    vals = [i+1, name, tier_label, score, msgs, tags, active, li, tw, ig, fb, web, role, snippet]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = DATA_FONT
        c.fill = fill
        c.border = THIN_BORDER
        c.alignment = Alignment(vertical="center", wrap_text=(col >= 6))
        if col in (8, 9, 10, 11, 12) and val:
            c.font = LINK_FONT

out = "_outputs/Key_Contacts_Social_Profiles.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Rows: {len(contacts)}")
